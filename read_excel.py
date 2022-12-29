import mmap
from io import BytesIO
import openpyxl
import xlrd
from xlrd.sheet import Sheet
from fastapi import UploadFile, File
from sqlalchemy.dialects.sqlite import insert
from sqlalchemy.orm import Session
from app.table_models import OneSiteUser, AwardsUser


def make_names_uniform(name: str) -> str:
    if "," in name:
        return " ".join(reversed(name.split(",")))
    else:
        return name


def mmap_io(f: UploadFile = File(...)):
    with mmap.mmap(f.file.fileno(), length=0, access=mmap.ACCESS_READ) as mmap_obj:
        text = mmap_obj.read()
        return text


def upload_xls_worksheet(db: Session, ws: Sheet):
    curr_row = 3
    num_rows = ws.nrows - 1
    while curr_row < num_rows:
        curr_row += 1
        row = ws.row_values(curr_row)
        user = dict(
            username="-".join([row[4], row[3]]),
            property=row[0],
            resh_id=row[1],
            lease_id=row[2],
            bldg_unit=row[3],
            name=make_names_uniform(row[4]),
            phone_number=row[5],
            email=row[6],
            status=row[7],
            move_in_out=row[8],
            code_description=row[9],
            total_prepaid=row[10],
            total_delinquent=row[11],
            d=row[12],
            o=row[13],
            net_balance=row[14],
            current=row[15],
            days_30=row[16],
            days_60=row[17],
            days_90_plus=row[18],
            prorate_credit=row[19],
            deposits_held=row[20],
            outstanding_deposit=row[21],
            late=row[22],
            nsf=row[23],
            delinquency_comment=row[24],
            comment_date=row[25],
            leasing_agent=row[26],
        )
        stmt = insert(OneSiteUser).values(user)
        upsert_stmt = stmt.on_conflict_do_update(
            index_elements=[OneSiteUser.username],
            set_=user,
        )
        db.execute(upsert_stmt)
    db.commit()


def read_xls(db: Session, file: UploadFile = File(...)):
    workbook = xlrd.open_workbook(file_contents=mmap_io(file), formatting_info=True)
    all_ws = workbook.sheet_by_name("All")
    curr_ws = workbook.sheet_by_name("Curr")
    upload_xls_worksheet(db, curr_ws)
    upload_xls_worksheet(db, all_ws)


def read_xlsx(db: Session, file: UploadFile = File(...)):
    workbook = openpyxl.load_workbook(filename=BytesIO(file.file.read()))
    worksheet = workbook.active
    num_rows = worksheet.max_row
    rows = worksheet.iter_rows(13, num_rows, 0, 43, values_only=True)
    values = [row for row in rows]
    for row in values:
        if row[0] is None:
            break
        else:
            a_user = dict(
                program=row[0],
                name=make_names_uniform(row[1]),
                rent_arrears_plan_form_date=row[2],
                total_client_arrears_90_plus_days_old=row[3],
                arrears_plan=row[4],
                erap_app_submitted=row[5],
                erap_app_date=row[6],
                erap_confirmation_number=row[7],
                erap_app_status=row[8],
                erap_status_date=row[9],
                erap_payment_received=row[10],
                erap_amount_received=row[11],
                erap_denied=row[12],
                osd_app_submitted=row[13],
                osd_app_date=row[14],
                osd_confirmation_number=row[15],
                osd_app_status=row[16],
                osd_status_date=row[17],
                osd_payment_received=row[18],
                osd_payment_amount=row[19],
                homebase_linkages=row[20],
                homebase_provider=row[21],
                ssi=row[22],
                ssi_rep_payee=row[23],
                case_manager_conference=row[24],
                cmc_date=row[25],
                rep_payee=row[26],
                voluntary_up_rep_payee=row[27],
                type_of_entitlement_issue=row[28],
                cmc_outcome=row[29],
                pm_case_conference=row[30],
                pm_cc_date=row[31],
                pl_cc_outcome=row[32],
                ecc=row[33],
                ecc_date=row[34],
                ecc_outcome=row[35],
                housing_court_action=row[36],
                housing_counsel_referral_date=row[37],
                date_letter_595_discharge=row[38],
                notice_595_housing_action=row[39],
                payment_plan_start_date=row[40],
                repayment_plan_amount=row[41],
                rent_arrears_plan_note=row[42],
                username="-".join([row[1], row[0], row[4]])
            )
            stmt = insert(AwardsUser).values(a_user)
            upsert_stmt = stmt.on_conflict_do_update(
                index_elements=[AwardsUser.username],
                set_=a_user
            )
            db.execute(upsert_stmt)
        db.commit()


if __name__ == '__main__':
    import requests
    import pandas as pd

    result = requests.get("http://127.0.0.1:8000/read-onesite-users/").json()
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_colwidth', None)
    answer = [entry["name"] for entry in result]
    df = pd.Series(answer)
    print(df)
