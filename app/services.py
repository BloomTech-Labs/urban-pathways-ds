import os
import pandas as pd
import openpyxl
import xlrd
from thefuzz import process, fuzz
from xlrd.sheet import Sheet
from app.tools import make_names_uniform, remove_multiple_periods, mmap_io


def upload_ws(worksheet: Sheet) -> pd.DataFrame:
    curr_row = 3
    num_rows = worksheet.nrows - 1
    all_users = []
    while curr_row < num_rows:
        curr_row += 1
        row = worksheet.row_values(curr_row)
        o_user = dict(
            property=row[0],
            resh_id=row[1],
            lease_id=row[2],
            bldg_unit=row[3],
            name=make_names_uniform(row[4]),
            phone_number=row[5],
            email=row[6],
            status=row[7],
            move_in_out=row[8],
            code_description=row[9],
            total_prepaid=row[10],
            total_delinquent=row[11],
            d=row[12],
            o=row[13],
            net_balance=row[14],
            current=row[15],
            days_30=row[16],
            days_60=row[17],
            days_90_plus=row[18],
            prorate_credit=row[19],
            deposits_held=row[20],
            outstanding_deposit=row[21],
            late=row[22],
            nsf=row[23],
            delinquency_comment=row[24],
            comment_date=row[25],
            leasing_agent=row[26],
        )
        all_users.append(o_user)
    return pd.DataFrame(all_users)


def read_xls(file):
    workbook = xlrd.open_workbook(file_contents=mmap_io(file), formatting_info=True)
    all_ws = workbook.sheet_by_name("All")
    curr_ws = workbook.sheet_by_name("Curr")
    all_df = upload_ws(all_ws)
    curr_df = upload_ws(curr_ws)
    return pd.concat([all_df, curr_df])


def read_xlsx(file):
    workbook = openpyxl.load_workbook(filename=file)
    worksheet = workbook.active
    num_rows = worksheet.max_row
    cleaned_sheet = worksheet.iter_rows(13, num_rows, 0, 43, values_only=True)
    values = [row for row in cleaned_sheet]
    all_users = []
    for row in values:
        if row[0] is None:
            break
        else:
            a_user = dict(
                program=row[0],
                name=make_names_uniform(row[1]),
                rent_arrears_plan_form_date=row[2],
                total_client_arrears_90_plus_days_old=remove_multiple_periods(row[3]),
                arrears_plan=row[4],
                erap_app_submitted=row[5],
                erap_app_date=row[6],
                erap_confirmation_number=row[7],
                erap_app_status=row[8],
                erap_status_date=row[9],
                erap_payment_received=row[10],
                erap_amount_received=row[11],
                erap_denied=row[12],
                osd_app_submitted=row[13],
                osd_app_date=row[14],
                osd_confirmation_number=row[15],
                osd_app_status=row[16],
                osd_status_date=row[17],
                osd_payment_received=row[18],
                osd_payment_amount=row[19],
                homebase_linkages=row[20],
                homebase_provider=row[21],
                ssi=row[22],
                ssi_rep_payee=row[23],
                case_manager_conference=row[24],
                cmc_date=row[25],
                rep_payee=row[26],
                voluntary_up_rep_payee=row[27],
                type_of_entitlement_issue=row[28],
                cmc_outcome=row[29],
                pm_case_conference=row[30],
                pm_cc_date=row[31],
                pl_cc_outcome=row[32],
                ecc=row[33],
                ecc_date=row[34],
                ecc_outcome=row[35],
                housing_court_action=row[36],
                housing_counsel_referral_date=row[37],
                date_letter_595_discharge=row[38],
                notice_595_housing_action=row[39],
                payment_plan_start_date=row[40],
                repayment_plan_amount=row[41],
                rent_arrears_plan_note=row[42],
            )
            all_users.append(a_user)
    return pd.DataFrame(all_users)


def fuzzy_merge_export(df1, df2, date, threshold=90, limit=1):
    s = df2["name"].tolist()
    m = df1["name"].apply(
        lambda x: process.extract(
            x,
            s,
            limit=limit,
            scorer=fuzz.token_sort_ratio
        )
    )
    df1["matches"] = m
    m2 = df1["matches"].apply(
        lambda x: ", ".join([i[0] for i in x if i[1] >= threshold])
    )
    df1["matches"] = m2
    matched_df = df1[df1["matches"] != ""]
    no_match_df = df1.loc[df1["matches"] == ""]
    matched_fp = os.path.join(
        os.path.abspath("app"),
        "csv",
        f"matched_clients_{date}.csv"
    )
    without_matches_fp = os.path.join(
        os.path.abspath("app"),
        "csv",
        f"clients_without_matches_{date}.csv"
    )
    matched_df.to_csv(matched_fp, index=False)
    no_match_df.to_csv(without_matches_fp, index=False)
